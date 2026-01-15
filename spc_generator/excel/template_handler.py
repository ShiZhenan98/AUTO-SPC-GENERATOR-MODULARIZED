"""模板处理器"""

from openpyxl import load_workbook, Workbook
from typing import Optional
from ..utils.date_utils import DateUtils
from ..config.constants import CELL_ADDRESSES


class TemplateHandler:
    """模板处理器"""
    
    def __init__(self):
        self.date_utils = DateUtils()
    
    def load_template(self, template_path: str) -> Workbook:
        """加载模板"""
        return load_workbook(template_path)
    
    def fill_basic_info(
        self,
        ws,
        workshop_name: str,
        product_model: str,
        process: str,
        theory: str,
        inspection_item: str,
        equipment_no: str,
        tolerance: object,
        year: int,
        month: int
    ):
        """填充基本信息"""
        # 车间信息
        ws[CELL_ADDRESSES['workshop']] = workshop_name
        
        # 产品型号
        ws[CELL_ADDRESSES['product_model']] = product_model
        ws[CELL_ADDRESSES['product_model_2']] = product_model
        
        # 工序信息
        ws[CELL_ADDRESSES['process']] = process if process else ""
        
        # 理论值
        ws[CELL_ADDRESSES['theory']] = theory if theory else ""
        
        # 公差
        if tolerance.usl is not None:
            ws[CELL_ADDRESSES['usl']] = tolerance.usl
        else:
            ws[CELL_ADDRESSES['usl']] = None
        
        if tolerance.lsl is not None:
            ws[CELL_ADDRESSES['lsl']] = tolerance.lsl
        else:
            ws[CELL_ADDRESSES['lsl']] = None
        
        # 检测项目
        ws[CELL_ADDRESSES['inspection_item']] = inspection_item if inspection_item else ""
        
        # 设备编号
        ws[CELL_ADDRESSES['equipment_no']] = equipment_no if equipment_no else ""
        
        # 生成并填充日期
        month_dates = self.date_utils.generate_month_dates(year, month, 25)
        for i, date_str in enumerate(month_dates):
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(3 + i)
            cell_address = f"{col_letter}{CELL_ADDRESSES['date_row']}"
            ws[cell_address] = date_str
