"""公式恢复器"""

from typing import Optional
from ..models.tolerance import Tolerance
from ..config.constants import CELL_ADDRESSES


class FormulaRestorer:
    """公式恢复器"""
    
    def restore_formulas(
        self,
        worksheet,
        tolerance: Tolerance,
        use_reference_range: bool = False,
        ref_center: Optional[float] = None
    ):
        """
        恢复原模板中的关键公式，根据单边/双边公差动态设置公式
        
        Args:
            worksheet: 工作表对象
            tolerance: 公差信息
            use_reference_range: 是否使用参考分布范围模式
            ref_center: 参考中心（用于参考分布范围模式）
        """
        try:
            # 检测公差类型并动态设置AA29的公式
            if tolerance.usl is not None and tolerance.lsl is not None:
                aa29_formula = '=MIN(X29,Y29)'
            elif tolerance.usl is not None:
                aa29_formula = '=X29'
            elif tolerance.lsl is not None:
                aa29_formula = '=Y29'
            else:
                aa29_formula = 0
            
            worksheet['AA29'] = aa29_formula
            
            # 如果使用参考分布范围，E4已经是数值，不恢复公式
            if not use_reference_range or ref_center is None:
                worksheet[CELL_ADDRESSES['center']] = '=(Q2+Q3)/2'
                if use_reference_range and ref_center is not None:
                    worksheet[CELL_ADDRESSES['center_reference']] = ref_center
            else:
                worksheet[CELL_ADDRESSES['center']] = ref_center
                worksheet[CELL_ADDRESSES['center_reference']] = ref_center
            
            # 恢复其他公式
            worksheet['H4'] = '=E17/2.326'
            worksheet['J4'] = '=E4+H4*3'
            worksheet['L4'] = '=E4-H4*3'
            worksheet['N4'] = '=E4+H4'
            worksheet['P4'] = '=E4+H4*2'
            worksheet['R4'] = '=E4-H4'
            worksheet['T4'] = '=E4-H4*2'
            worksheet['E17'] = '=AVERAGE(C38:AA38)'
            worksheet['L17'] = '=Z19*E17'
            worksheet['U17'] = '=0*E17'
            worksheet['X29'] = '=ABS((Q2-AVERAGE(C32:AA36))/(3*H4))'
            worksheet['Y29'] = '=ABS(AVERAGE(C32:AA36)-Q3)/(3*H4)'
            
            # 平均值行公式 (C37:AA37)
            for col in range(3, 29):  # C到AA列
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(col)
                worksheet[f'{col_letter}37'] = f'=AVERAGE({col_letter}32:{col_letter}36)'
            
            # 极差行公式 (C38:AA38)
            for col in range(3, 29):
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(col)
                worksheet[f'{col_letter}38'] = f'=MAX({col_letter}32:{col_letter}36)-MIN({col_letter}32:{col_letter}36)'
            
        except Exception as e:
            print(f"恢复公式时出错: {e}")
