"""工作表写入器"""

from typing import List, Dict
import numpy as np
from openpyxl.utils import get_column_letter
from ..models.task import Task
from ..models.tolerance import Tolerance
from ..models.control_limits import ControlLimits
from ..models.spc_data import SPCData
from ..utils.date_utils import DateUtils
from ..utils.validation_utils import ValidationUtils
from ..config.constants import CELL_ADDRESSES


class WorksheetWriter:
    """工作表写入器"""
    
    def __init__(self):
        self.date_utils = DateUtils()
        self.validator = ValidationUtils()
    
    def fill_measurement_data(
        self,
        worksheet,
        spc_data: SPCData
    ):
        """
        填充测量数据
        
        Args:
            worksheet: 工作表对象
            spc_data: SPC数据对象
        """
        # 写入舍入后的测量数据 (C32:AA36)
        for row_idx in range(5):
            for col_idx in range(25):
                try:
                    if spc_data.rounded_measurement_data:
                        value = spc_data.rounded_measurement_data[row_idx][col_idx]
                    else:
                        value = spc_data.measurement_data[row_idx][col_idx]
                    
                    if value is not None:
                        worksheet.cell(row=32+row_idx, column=3+col_idx).value = value
                except Exception:
                    pass
        
        # 更新平均值和极差行（使用舍入后的数据重新计算）
        for col_idx in range(25):
            try:
                if spc_data.rounded_measurement_data:
                    subgroup = [
                        spc_data.rounded_measurement_data[row_idx][col_idx]
                        for row_idx in range(5)
                    ]
                else:
                    subgroup = [
                        spc_data.measurement_data[row_idx][col_idx]
                        for row_idx in range(5)
                    ]
                
                # 计算平均值和极差
                avg = np.mean(subgroup)
                range_val = max(subgroup) - min(subgroup)
                
                # 平均值小数位数：根据分辨率确定
                avg_decimal_places = spc_data.max_decimal_places + 1
                
                worksheet.cell(row=37, column=3+col_idx).value = self.validator.format_value(
                    avg, avg_decimal_places
                )
                worksheet.cell(row=38, column=3+col_idx).value = self.validator.format_value(
                    range_val, 3
                )
            except Exception:
                pass
    
    def fill_approver_info(
        self,
        worksheet,
        approver_info: Dict[str, str],
        year: int,
        month: int
    ):
        """
        填充审批人信息
        
        Args:
            worksheet: 工作表对象
            approver_info: 审批人信息字典
            year: 年份
            month: 月份
        """
        month_end_date = self.date_utils.format_month_end_date(year, month)
        
        if approver_info.get('preparer'):
            try:
                worksheet[CELL_ADDRESSES['preparer']] = f"{approver_info['preparer']} {month_end_date}"
            except:
                pass
        
        if approver_info.get('auditor'):
            try:
                worksheet[CELL_ADDRESSES['auditor']] = f"{approver_info['auditor']} {month_end_date}"
            except:
                pass
        
        if approver_info.get('approver'):
            try:
                worksheet[CELL_ADDRESSES['approver']] = f"{approver_info['approver']} {month_end_date}"
            except:
                pass
    
    def set_reference_center(
        self,
        worksheet,
        ref_center: float,
        use_reference_range: bool
    ):
        """
        设置参考中心
        
        Args:
            worksheet: 工作表对象
            ref_center: 参考中心值
            use_reference_range: 是否使用参考范围模式
        """
        if use_reference_range:
            worksheet[CELL_ADDRESSES['center']] = ref_center
            worksheet[CELL_ADDRESSES['center_reference']] = ref_center
