"""Excel计划文件读取器"""

import re
from typing import List, Dict
from openpyxl import load_workbook
from ..models.task import Task
from ..utils.validation_utils import ValidationUtils


class ExcelPlanReader:
    """Excel计划文件读取器"""
    
    def __init__(self):
        self.validator = ValidationUtils()
    
    def read_tasks(self, workbook_path: str) -> List[Task]:
        """
        读取所有任务
        
        Args:
            workbook_path: Excel文件路径
            
        Returns:
            任务列表
        """
        wb = load_workbook(workbook_path, data_only=True)
        ws = wb.active
        
        tasks = []
        for row in range(3, ws.max_row + 1):
            product_model = ws.cell(row=row, column=2).value  # B列
            if product_model is None or str(product_model).strip() == '':
                continue
            
            product_model_str = str(product_model).strip()
            if not product_model_str.startswith('NB'):
                continue
            
            # 读取任务信息
            task = Task(
                row_index=row,
                product_model=product_model_str,
                process=str(ws.cell(row=row, column=3).value).strip() if ws.cell(row=row, column=3).value else '',  # C列
                theory=str(ws.cell(row=row, column=4).value).strip() if ws.cell(row=row, column=4).value else '',  # D列
                reference_range=str(ws.cell(row=row, column=5).value).strip() if ws.cell(row=row, column=5).value else '',  # E列
                inspection_item=str(ws.cell(row=row, column=6).value).strip() if ws.cell(row=row, column=6).value else '',  # F列
                resolution=self.validator.safe_float_convert(ws.cell(row=row, column=7).value),  # G列
                equipment_no=str(ws.cell(row=row, column=8).value).strip() if ws.cell(row=row, column=8).value else '',  # H列
                target_cpk=self.validator.safe_float_convert(ws.cell(row=row, column=21).value) if ws.cell(row=row, column=21).value else 1.8,  # U列
                month_status={}
            )
            
            # 读取每个月份的状态 (I列到T列)
            for col in range(9, 21):  # I列(9)到T列(20)
                month_cell = ws.cell(row=row, column=col)
                month_value = month_cell.value
                
                if month_value is not None:
                    month_str = str(month_value).strip().upper()
                    if month_str == 'N':
                        # 计算月份索引: col-8 (因为I列是第1个月)
                        month_index = col - 8
                        month_name = f"{month_index}月"
                        task.month_status[month_name] = True
            
            tasks.append(task)
        
        wb.close()
        return tasks
    
    def read_approver_info(self, workbook_path: str) -> Dict[str, str]:
        """
        读取审批人信息
        
        Args:
            workbook_path: Excel文件路径
            
        Returns:
            审批人信息字典 {'preparer': '', 'auditor': '', 'approver': ''}
        """
        wb = load_workbook(workbook_path, data_only=True)
        ws = wb.active
        
        approver_info = self._get_approver_info_generic(ws)
        wb.close()
        return approver_info
    
    def read_month_names(self, workbook_path: str) -> Dict[int, str]:
        """
        读取月份名称映射
        
        Args:
            workbook_path: Excel文件路径
            
        Returns:
            月份名称映射 {列索引: 月份名称}
        """
        wb = load_workbook(workbook_path, data_only=True)
        ws = wb.active
        
        month_names = {}
        for col in range(9, 21):  # I列(9)到T列(20)
            month_cell = ws.cell(row=1, column=col)
            month_name = None
            if month_cell.value:
                month_name = str(month_cell.value).strip()
            
            if not month_name or month_name == '':
                month_num = col - 8  # I列是第1个月
                month_name = f"{month_num}月"
            elif month_name.upper() in ['N', 'Y']:
                month_num = col - 8
                month_name = f"{month_num}月"
            
            month_names[col] = month_name
        
        wb.close()
        return month_names
    
    def _get_approver_info_generic(self, ws) -> Dict[str, str]:
        """
        通用的审批人信息查找方法
        适用于各种SPC计划表格布局
        """
        try:
            approver_info = {
                'preparer': '',
                'auditor': '',
                'approver': ''
            }
            
            preparer_keywords = ['编制', '制表', '编制人', '制表人', 'preparer', 'author']
            auditor_keywords = ['审核', '审核人', 'auditor', 'checker']
            approver_keywords = ['批准', '批准人', '批准者', 'approver', 'approval']
            
            max_row = ws.max_row
            max_col = ws.max_column
            
            # 首先尝试从表格底部查找（最后10行）
            search_rows = min(10, max_row)
            start_row = max(1, max_row - search_rows + 1)
            
            for row in range(start_row, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell and cell.value:
                        cell_value = str(cell.value).strip()
                        
                        if not approver_info['preparer']:
                            for keyword in preparer_keywords:
                                if keyword in cell_value:
                                    name = self._extract_name_from_cell(cell_value, keyword, row, col, ws)
                                    if name:
                                        approver_info['preparer'] = name
                                        break
                        
                        if not approver_info['auditor']:
                            for keyword in auditor_keywords:
                                if keyword in cell_value:
                                    name = self._extract_name_from_cell(cell_value, keyword, row, col, ws)
                                    if name:
                                        approver_info['auditor'] = name
                                        break
                        
                        if not approver_info['approver']:
                            for keyword in approver_keywords:
                                if keyword in cell_value:
                                    name = self._extract_name_from_cell(cell_value, keyword, row, col, ws)
                                    if name:
                                        approver_info['approver'] = name
                                        break
            
            # 如果没找到，扩大搜索范围（整个表格）
            if not (approver_info['preparer'] and approver_info['auditor'] and approver_info['approver']):
                for row in range(1, max_row + 1):
                    if approver_info['preparer'] and approver_info['auditor'] and approver_info['approver']:
                        break
                    
                    for col in range(1, max_col + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell and cell.value:
                            cell_value = str(cell.value).strip()
                            
                            if not approver_info['preparer']:
                                for keyword in preparer_keywords:
                                    if keyword in cell_value:
                                        name = self._extract_name_from_cell(cell_value, keyword, row, col, ws)
                                        if name:
                                            approver_info['preparer'] = name
                                            break
                            
                            if not approver_info['auditor']:
                                for keyword in auditor_keywords:
                                    if keyword in cell_value:
                                        name = self._extract_name_from_cell(cell_value, keyword, row, col, ws)
                                        if name:
                                            approver_info['auditor'] = name
                                            break
                            
                            if not approver_info['approver']:
                                for keyword in approver_keywords:
                                    if keyword in cell_value:
                                        name = self._extract_name_from_cell(cell_value, keyword, row, col, ws)
                                        if name:
                                            approver_info['approver'] = name
                                            break
            
            # 如果仍然没有找到所有审批人，尝试从特定位置查找
            if not approver_info['preparer']:
                common_preparer_cells = ['C15', 'B15', 'D15', 'C10', 'B10']
                for cell_addr in common_preparer_cells:
                    try:
                        cell = ws[cell_addr]
                        if cell and cell.value:
                            cell_value = str(cell.value).strip()
                            if not re.match(r'^\d+$', cell_value) and 2 <= len(cell_value) <= 10:
                                approver_info['preparer'] = cell_value
                                break
                    except:
                        pass
            
            # 如果没有找到任何审批人，尝试从最后一行提取
            if not any(approver_info.values()):
                last_row_cells = []
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=max_row, column=col)
                    if cell and cell.value:
                        last_row_cells.append(str(cell.value).strip())
                
                for text in last_row_cells:
                    if re.search(r'[\u4e00-\u9fff]{2,4}', text):
                        cleaned = re.sub(r'[\d\s:：]', '', text)
                        if 2 <= len(cleaned) <= 4:
                            if not approver_info['preparer']:
                                approver_info['preparer'] = cleaned
                            elif not approver_info['auditor']:
                                approver_info['auditor'] = cleaned
                            elif not approver_info['approver']:
                                approver_info['approver'] = cleaned
            
            return approver_info
            
        except Exception as e:
            print(f"读取审批人信息时出错: {e}")
            return {'preparer': '', 'auditor': '', 'approver': ''}
    
    def _extract_name_from_cell(self, cell_value: str, keyword: str, row: int, col: int, ws) -> str:
        """从单元格中提取姓名"""
        try:
            patterns = [
                rf'{keyword}[:：]\s*([^\s:：]+)',
                rf'{keyword}\s+([^\s:：]+)',
                rf'{keyword}\s*([^\s:：]+)$',
            ]
            
            for pattern in patterns:
                match = re.search(pattern, cell_value)
                if match:
                    name = match.group(1).strip()
                    if self.validator.is_valid_name(name):
                        return name
            
            # 检查右侧单元格
            right_cell = ws.cell(row=row, column=col+1)
            if right_cell and right_cell.value:
                right_value = str(right_cell.value).strip()
                if self.validator.is_valid_name(right_value):
                    return right_value
            
            # 检查下方单元格
            below_cell = ws.cell(row=row+1, column=col)
            if below_cell and below_cell.value:
                below_value = str(below_cell.value).strip()
                if self.validator.is_valid_name(below_value):
                    return below_value
            
            return ''
            
        except Exception as e:
            return ''
