"""计划更新服务"""

import os
from typing import List, Dict
from openpyxl import load_workbook


class PlanUpdater:
    """计划更新服务"""
    
    def update_spc_plan_file(
        self,
        plan_file_path: str,
        results: List[Dict]
    ):
        """
        更新SPC推进计划文件
        
        Args:
            plan_file_path: 计划文件路径
            results: 结果列表，每个结果包含row_index, month_num, actual_cpk等
        """
        try:
            if not results:
                print("没有需要更新的结果")
                return
            
            # 加载原文件
            wb = load_workbook(plan_file_path)
            ws = wb.active
            
            # 实际cpk列（V列，第22列）
            actual_cpk_col = 22
            # 目标cpk列（U列，第21列）
            target_cpk_col = 21
            
            # 更新数据
            for result in results:
                row_idx = result['row_index']
                month_num = result['month_num']
                
                # 将N改为Y
                month_col = 8 + month_num  # I列是9，对应索引9
                ws.cell(row=row_idx, column=month_col).value = 'Y'
                
                # 填入实际cpk（V列，第22列）
                ws.cell(row=row_idx, column=actual_cpk_col).value = result['actual_cpk']
                
                # 更新目标cpk（U列，第21列）为调整后的值
                adjusted_target_cpk = result.get('adjusted_target_cpk')
                if adjusted_target_cpk is not None:
                    ws.cell(row=row_idx, column=target_cpk_col).value = adjusted_target_cpk
                    print(f"  更新行{row_idx}: {result['product_model']} - {result['process']}, "
                          f"目标cpk: {adjusted_target_cpk:.3f}, 实际cpk: {result['actual_cpk']:.3f}, "
                          f"难度: {result.get('difficulty', '未知')}")
                else:
                    print(f"  更新行{row_idx}: {result['product_model']} - {result['process']}, "
                          f"实际cpk: {result['actual_cpk']:.3f}, 难度: {result.get('difficulty', '未知')}")
            
            # 保存更新后的文件
            base_name = os.path.splitext(plan_file_path)[0]
            updated_filename = f"{base_name}_已更新.xlsx"
            wb.save(updated_filename)
            wb.close()
            
            print(f"\n已更新SPC推进计划文件: {updated_filename}")
        
        except Exception as e:
            print(f"更新SPC推进计划文件时出错: {e}")
            import traceback
            traceback.print_exc()
